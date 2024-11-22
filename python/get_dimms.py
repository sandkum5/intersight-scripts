#!/usr/bin/env python3
"""
    Intersight API calls using oAuth2
    Python Libraries Needed:
        pip install requests openpyxl python-dotenv

"""

import os
import sys
import json
import requests
from dotenv import load_dotenv, find_dotenv
from openpyxl.workbook import Workbook

load_dotenv(find_dotenv())

def get_token(client_id, client_secret):
    """ Get oAuth Token """
    token_url="https://intersight.com/iam/token"
    client_auth = requests.auth.HTTPBasicAuth(client_id, client_secret)
    post_data = {"grant_type": "client_credentials"}
    response = requests.post(url=token_url,
                            auth=client_auth,
                            data=post_data)
    if response.status_code != 200:
        print("Failed to obtain token from the OAuth 2.0 server", file=sys.stderr)
        sys.exit(1)
    print("Successfuly obtained a new token")
    token_json = response.json()
    return token_json["access_token"]

def get_api_data(token, client_id, client_secret, api_url):
    """ Get API Endpoint Data """
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(url=api_url, headers=headers)
    data = response.json()
    if	response.status_code == 401:
        print("Existing Token Expired. Generating a new one!")
        token = get_token(client_id, client_secret)
        get_api_data(token, client_id, client_secret, api_url)
    else:
        return data

if __name__ == '__main__':
    # Set variables
    client_id = os.getenv("ClientId")
    client_secret = os.getenv("ClientSecret")
    api_url = "https://intersight.com/api/v1/memory/Units?$expand=Ancestors($select=Name,Model,Dn,Serial)&$select=Ancestors,ArrayId,Bank,Capacity,ClassId,Clock,Description,Dn,FormFactor,Latency,Location,MemoryId,Model,Moid,ObjectType,OperState,PartNumber,Pid,Presence,Revision,Serial,Type,Vendor,Width"

    # Get oAuth Token
    token = get_token(client_id, client_secret)

    # Get Endpoint Data Count
    api_count_url = "https://intersight.com/api/v1/memory/Units?$count=True"
    count_data = get_api_data(token, client_id, client_secret, api_count_url)
    total_count = count_data["Count"]

    # Get API Endpoint Data
    increment = 1000
    skip = 0
    dimm_data = []
    while (skip <= total_count):
        api_url = f"https://intersight.com/api/v1/memory/Units?$expand=Ancestors($select=Name,Model,Dn,Serial)&$select=Ancestors,ArrayId,Bank,Capacity,ClassId,Clock,Description,Dn,FormFactor,Latency,Location,MemoryId,Model,Moid,ObjectType,OperState,PartNumber,Pid,Presence,Revision,Serial,Type,Vendor,Width&$top=1000&$skip={skip}"
        dimm_info = get_api_data(token, client_id, client_secret, api_url)
        dimm_data.extend(dimm_info["Results"])
        skip += increment

    # with open('dimms.json', 'w') as f:
    #     f.write(json.dumps(dimm_data))


# Create Excel File
datafile = 'DIMMs.xlsx'

# Create Workbook
workbook = Workbook()
workbook.save(datafile)

sheet = workbook.active 
sheet.title = 'DIMMs'

# Write Sheet Headers
headers = [ "Ancestors.Blade_Name", "Ancestors.Blade_Model", "Ancestors.Blade_Serial", "ArrayId", "Bank","Capacity", "ClassId", "Clock", "Description", "Dn", "FormFactor", "Latency", "Location", "MemoryId", "Model", "Moid", "ObjectType", "OperState", "PartNumber", "Pid", "Presence", "Revision", "Serial", "Type", "Vendor", "Width" ]

for column, value in enumerate(headers, start=1):
    sheet.cell(row=1, column=column, value=value)

# Write Sheet Data

for row, value in enumerate(dimm_data, start=2):
    sheet.cell(row=row,column=1, value=value["Ancestors"][2]["Name"])
    sheet.cell(row=row,column=2, value=value["Ancestors"][2]["Model"])
    sheet.cell(row=row,column=3, value=value["Ancestors"][2]["Serial"])
    sheet.cell(row=row,column=4, value=value["ArrayId"])
    sheet.cell(row=row,column=5, value=value["Bank"])
    sheet.cell(row=row,column=6, value=value["Capacity"])
    sheet.cell(row=row,column=7, value=value["ClassId"])
    sheet.cell(row=row,column=8, value=value["Clock"])
    sheet.cell(row=row,column=9, value=value["Description"])
    sheet.cell(row=row,column=10, value=value["Dn"])
    sheet.cell(row=row,column=11, value=value["FormFactor"])
    sheet.cell(row=row,column=12, value=value["Latency"])
    sheet.cell(row=row,column=13, value=value["Location"])
    sheet.cell(row=row,column=14, value=value["MemoryId"])
    sheet.cell(row=row,column=15, value=value["Model"])
    sheet.cell(row=row,column=16, value=value["Moid"])
    sheet.cell(row=row,column=17, value=value["ObjectType"])
    sheet.cell(row=row,column=18, value=value["OperState"])
    sheet.cell(row=row,column=19, value=value["PartNumber"])
    sheet.cell(row=row,column=20, value=value["Pid"])
    sheet.cell(row=row,column=21, value=value["Presence"])
    sheet.cell(row=row,column=22, value=value["Revision"])
    sheet.cell(row=row,column=23, value=value["Serial"])
    sheet.cell(row=row,column=24, value=value["Type"])
    sheet.cell(row=row,column=25, value=value["Vendor"])
    sheet.cell(row=row,column=26, value=value["Width"])

workbook.save(datafile)
workbook.close()
